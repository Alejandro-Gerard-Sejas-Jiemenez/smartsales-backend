import re
import logging
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from io import BytesIO

from decimal import Decimal
from datetime import datetime, date

# Importar modelos
from apps.venta_transacciones.models import Venta, DetalleVenta
from apps.catalogo.models import Producto, Categoria, Cliente

# Importar librerías de reportes
try:
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ===================================================================
# 1. EL "INTÉRPRETE" DE PROMPTS (CU-12)
# ===================================================================

# --- (Las funciones 'parse_date_range_from_prompt', 'parse_format_from_prompt', 
#      y 'parse_grouping_from_prompt' están correctas, no se modifican) ---

def parse_date_range_from_prompt(text):
    text = text.lower()
    today = timezone.now().date()
    match = re.search(r'del\s+(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})', text)
    if match:
        try:
            start_date = datetime.strptime(match.group(1), '%d/%m/%Y').date()
            end_date = datetime.strptime(match.group(2), '%d/%m/%Y').date()
            return start_date, end_date
        except ValueError:
            pass
    match = re.search(r'(últimos|ultimos)\s+(\d+)\s+d(í|i)as', text)
    if match:
        days = int(match.group(2))
        start_date = today - timedelta(days=days - 1)
        return start_date, today
    if re.search(r'este\s+mes', text) or re.search(r'(últimos|ultimos)\s+30\s+d(í|i)as', text):
        start_date = today.replace(day=1)
        return start_date, today
    if re.search(r'mes\s+pasado', text):
        first_of_this_month = today.replace(day=1)
        last_of_last_month = first_of_this_month - timedelta(days=1)
        first_of_last_month = last_of_last_month.replace(day=1)
        return first_of_last_month, last_of_last_month
    months = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    for month_name, month_num in months.items():
        if month_name in text:
            year = today.year
            start_date = datetime(year, month_num, 1).date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            return start_date, end_date
    return None, None

def parse_format_from_prompt(text):
    text = text.lower()
    if re.search(r'excel|xlsx', text):
        return 'excel'
    return 'pdf'

def parse_grouping_from_prompt(text):
    text = text.lower()
    if re.search(r'(agrupado|agrupar)\s+por\s+producto', text):
        return 'producto'
    if re.search(r'(agrupado|agrupar)\s+por\s+cliente', text):
        return 'cliente'
    if re.search(r'(agrupado|agrupar)\s+por\s+categor(í|i)a', text):
        return 'categoria'
    return None

def generate_dynamic_report(prompt):
    # 1. Interpretar el Prompt
    date_from, date_to = parse_date_range_from_prompt(prompt)
    output_format = parse_format_from_prompt(prompt)
    group_by = parse_grouping_from_prompt(prompt)
    
    # 2. Construir la Consulta Base
    query = DetalleVenta.objects.select_related('venta', 'producto', 'venta__cliente__usuario', 'producto__categoria')
    
    if date_from and date_to:
        query = query.filter(venta__fecha_venta__date__range=[date_from, date_to])

    # 3. Aplicar Agrupación y Calcular Estadísticas
    
    if group_by == 'producto':
        data = query.values('producto__nombre') \
                    .annotate(
                        cantidad_total=Sum('cantidad'),
                        monto_total=Sum('subtotal')
                    ).order_by('-monto_total')
        headers = ["Producto", "Cantidad Total Vendida", "Monto Total (Bs)"]
        title = "Reporte de Ventas por Producto"
        # stats provisional, se convertirán a tipos básicos luego
        stats = {
            'total_ventas': data.count(),
            'cantidad_total': data.aggregate(Sum('cantidad_total'))['cantidad_total__sum'] or 0,
            'monto_total': data.aggregate(Sum('monto_total'))['monto_total__sum'] or 0,
        }
        
    elif group_by == 'cliente':
        data = query.values('venta__cliente__usuario__correo') \
                    .annotate(
                        cantidad_total=Sum('cantidad'),
                        monto_total=Sum('subtotal'),
                        cantidad_ventas=Count('venta__id', distinct=True)
                    ).order_by('-monto_total')
        headers = ["Cliente (Correo)", "Cantidad de Ventas", "Monto Total (Bs)"]
        title = "Reporte de Ventas por Cliente"
        stats = {
            'total_ventas': data.aggregate(Sum('cantidad_ventas'))['cantidad_ventas__sum'] or 0,
            'cantidad_total': data.aggregate(Sum('cantidad_total'))['cantidad_total__sum'] or 0,
            'monto_total': data.aggregate(Sum('monto_total'))['monto_total__sum'] or 0,
        }

    elif group_by == 'categoria':
        data = query.values('producto__categoria__nombre') \
                    .annotate(
                        cantidad_total=Sum('cantidad'),
                        monto_total=Sum('subtotal')
                    ).order_by('-monto_total')
        headers = ["Categoría", "Cantidad Total Vendida", "Monto Total (Bs)"]
        title = "Reporte de Ventas por Categoría"
        stats = {
            'total_ventas': data.count(),
            'cantidad_total': data.aggregate(Sum('cantidad_total'))['cantidad_total__sum'] or 0,
            'monto_total': data.aggregate(Sum('monto_total'))['monto_total__sum'] or 0,
        }
    
    else:
        # Reporte simple (lista de ventas)
        group_by = 'general'
        data = Venta.objects.select_related('cliente__usuario').order_by('-fecha_venta')
        if date_from and date_to:
            data = data.filter(fecha_venta__date__range=[date_from, date_to])
        headers = ["ID Venta", "Fecha", "Cliente", "Método", "Total (Bs)"]
        title = "Reporte General de Ventas"
        stats = {
            'total_ventas': data.count(),
            'monto_total': data.aggregate(Sum('total'))['total__sum'] or 0,
            'cantidad_total': DetalleVenta.objects.filter(venta__in=data).aggregate(Sum('cantidad'))['cantidad__sum'] or 0,
        }

    # --- Normalizar stats a tipos básicos para evitar Decimals/None en excel ---
    try:
        stats['total_ventas'] = int(stats.get('total_ventas') or 0)
    except Exception:
        try:
            stats['total_ventas'] = int(float(stats.get('total_ventas') or 0))
        except Exception:
            stats['total_ventas'] = 0

    try:
        stats['cantidad_total'] = int(stats.get('cantidad_total') or 0)
    except Exception:
        try:
            stats['cantidad_total'] = int(float(stats.get('cantidad_total') or 0))
        except Exception:
            stats['cantidad_total'] = 0

    try:
        stats['monto_total'] = float(stats.get('monto_total') or 0.0)
    except Exception:
        stats['monto_total'] = 0.0

    # 4. Validar si hay datos
    # Para agrupaciones (values/annotate) materializamos en lista para comprobar vacío de forma segura
    if group_by == 'general':
        if not data.exists():
            raise Venta.DoesNotExist("No se encontraron datos para el reporte solicitado.")
        # dejamos `data` como queryset para el PDF (no lo materializamos aquí)
    else:
        # materializar los resultados ya que son ValuesQuerySet (diccionarios)
        data_list = list(data)
        if not data_list:
            raise Venta.DoesNotExist("No se encontraron datos para el reporte solicitado.")
        # reasignamos `data` a la lista para uso en Excel (y evitar evaluación perezosa)
        data = data_list

    # 5. Crear la respuesta HTTP
    if output_format == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="reporte_smart_sales.xlsx"'},
        )

        # Llamada protegida al generador de Excel: imprimimos traceback y propagamos la excepción
        try:
            # materializar también el queryset 'general' si es que es el caso, para estabilidad
            if group_by == 'general':
                data_for_excel = list(data)  # venta objects list
            else:
                data_for_excel = data  # ya es lista

            return generate_excel_report(data_for_excel, title, headers, stats, group_by, response)

        except Exception as e:
            # Registramos/mostramos el traceback aquí para que aparezca en la consola de Django
            import traceback
            traceback.print_exc()
            # Re-lanzamos la excepción para que tu ViewSet la capture y muestre el error también
            raise

    else: # PDF es el default (no modifico la lógica del PDF)
        response = HttpResponse(
            content_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="reporte_smart_sales.pdf"'},
        )
        return generate_pdf_report(data, title, headers, stats, group_by, response)


# ===================================================================
# 2. GENERADORES DE ARCHIVOS
# ===================================================================
logger = logging.getLogger(__name__)
STYLES = getSampleStyleSheet()
TITLE_STYLE = ParagraphStyle('CustomTitle', parent=STYLES['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
SUBTITLE_STYLE = ParagraphStyle('Subtitle', parent=STYLES['Normal'], fontSize=9, textColor=colors.HexColor('#6B7280'), alignment=TA_CENTER, spaceAfter=12)
HEADER_STYLE = ParagraphStyle('Header', parent=STYLES['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)

def _build_pdf_header(elements):
    elements.append(Paragraph('📊 REPORTE DE VENTAS (SmartSales365)', TITLE_STYLE))
    elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", SUBTITLE_STYLE))
    elements.append(Spacer(1, 0.2*inch))

def _build_pdf_stats_cards(elements, stats):
    stats_data = [
        ['Total Ventas/Filas', 'Cantidad Total Items', 'Monto Total (Bs)'],
        [
            f"{stats['total_ventas']}",
            f"{stats['cantidad_total']:.0f}",
            f"{stats['monto_total']:.2f}"
        ]
    ]
    stats_table = Table(stats_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F3F4F6')),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10)
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))


def generate_pdf_report(data, title, headers, stats, group_by, response):
    if not REPORTLAB_AVAILABLE:
        raise ImportError("ReportLab no está instalado.")
        
    buffer = BytesIO()
    # Generamos el PDF en el buffer
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    _build_pdf_header(elements)
    _build_pdf_stats_cards(elements, stats)
    
    elements.append(Paragraph(f'<b>📋 {title}</b>', HEADER_STYLE))
    elements.append(Spacer(1, 10))
    
    table_data = [headers]
    
    if group_by == 'general':
        col_widths = [1*inch, 1.5*inch, 2.5*inch, 1.5*inch, 1.5*inch]
        for venta in data[:100]:
            fecha_str = venta.fecha_venta.strftime('%d/%m/%Y %H:%M') if venta.fecha_venta else 'N/A'
            cliente = venta.cliente.usuario.correo if getattr(venta, 'cliente', None) else 'N/A'
            table_data.append([
                getattr(venta, 'id', 'N/A'),
                fecha_str,
                cliente,
                getattr(venta, 'metodo_entrada', 'N/A'),
                f"{getattr(venta, 'total', 0):.2f} Bs"
            ])
    elif group_by == 'producto':
        col_widths = [3*inch, 2*inch, 2*inch]
        for item in data[:100]:
            table_data.append([
                item.get('producto__nombre'),
                item.get('cantidad_total') or 0,
                f"{(item.get('monto_total') or 0):.2f} Bs"
            ])
    elif group_by == 'cliente':
        col_widths = [3*inch, 2*inch, 2*inch]
        for item in data[:100]:
            table_data.append([
                item.get('venta__cliente__usuario__correo'),
                item.get('cantidad_ventas') or 0,
                f"{(item.get('monto_total') or 0):.2f} Bs"
            ])
    elif group_by == 'categoria':
        col_widths = [3*inch, 2*inch, 2*inch]
        for item in data[:100]:
            table_data.append([
                item.get('producto__categoria__nombre'),
                item.get('cantidad_total') or 0,
                f"{(item.get('monto_total') or 0):.2f} Bs"
            ])
    else:
        # fallback si viene algo raro
        col_widths = [2*inch] * len(headers)
    
    detail_table = Table(table_data, colWidths=col_widths, hAlign='LEFT')
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB'))
    ]))
    elements.append(detail_table)
    
    # Construimos el documento en el buffer
    doc.build(elements)
    buffer.seek(0)

    # Escribimos el buffer en la respuesta HTTP y devolvemos
    response.write(buffer.getvalue())
    buffer.close()
    return response


def generate_excel_report(data, title, headers, stats, group_by, response):
    """
    Genera un reporte Excel robusto con manejo completo de errores y logging.
    """
    import traceback
    import logging
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    logger = logging.getLogger(__name__)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Estilos
        border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')

        # === Cabecera ===
        ws['A1'] = f'📊 {title.upper()}'
        ws['A1'].font = Font(bold=True, size=20, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 35

        ws['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = center_align

        # === KPIs ===
        ws['A4'] = 'Total Ventas'
        ws['B4'] = stats.get('total_ventas', 0)
        ws['A5'] = 'Cantidad Total'
        ws['B5'] = stats.get('cantidad_total', 0)
        ws['A6'] = 'Monto Total (Bs.)'
        ws['B6'] = stats.get('monto_total', 0)
        for r in range(4, 7):
            ws[f'A{r}'].font = Font(bold=True)
            ws[f'A{r}'].alignment = center_align
            ws[f'B{r}'].alignment = center_align

        # === Encabezado tabla ===
        start_row = 8
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # === Contenido ===
        from decimal import Decimal
        from datetime import datetime, date

        def safe_val(v):
            if v is None:
                return ''
            if isinstance(v, Decimal):
                return float(v)
            if isinstance(v, (datetime, date)):
                return v.replace(tzinfo=None) if hasattr(v, 'tzinfo') else v
            return str(v)

        row = start_row + 1
        for item in data:
            try:
                if group_by == 'general':
                    ws.cell(row=row, column=1, value=safe_val(getattr(item, 'id', None)))
                    ws.cell(row=row, column=2, value=safe_val(getattr(item, 'fecha_venta', None)))
                    cliente = getattr(item.cliente.usuario, 'correo', 'N/A') if getattr(item, 'cliente', None) else 'N/A'
                    ws.cell(row=row, column=3, value=safe_val(cliente))
                    ws.cell(row=row, column=4, value=safe_val(getattr(item, 'metodo_entrada', 'N/A')))
                    ws.cell(row=row, column=5, value=safe_val(getattr(item, 'total', 0)))
                elif group_by == 'producto':
                    ws.cell(row=row, column=1, value=safe_val(item.get('producto__nombre')))
                    ws.cell(row=row, column=2, value=safe_val(item.get('cantidad_total')))
                    ws.cell(row=row, column=3, value=safe_val(item.get('monto_total')))
                elif group_by == 'cliente':
                    ws.cell(row=row, column=1, value=safe_val(item.get('venta__cliente__usuario__correo')))
                    ws.cell(row=row, column=2, value=safe_val(item.get('cantidad_ventas')))
                    ws.cell(row=row, column=3, value=safe_val(item.get('monto_total')))
                elif group_by == 'categoria':
                    ws.cell(row=row, column=1, value=safe_val(item.get('producto__categoria__nombre')))
                    ws.cell(row=row, column=2, value=safe_val(item.get('cantidad_total')))
                    ws.cell(row=row, column=3, value=safe_val(item.get('monto_total')))
                row += 1
            except Exception as row_err:
                logger.warning(f"Fila con error {row}: {row_err}")
                traceback.print_exc()
                row += 1
                continue

        # === Ajuste de ancho ===
        for i, column_cells in enumerate(ws.columns, 1):
            # Filtra celdas fusionadas y valores None
            valid_cells = [cell for cell in column_cells if cell.value is not None and not isinstance(cell, type(ws['A1']))]
            if not valid_cells:
                continue
            max_length = max(len(str(cell.value)) for cell in valid_cells)
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max_length + 2

        # === Guardar ===
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        try:
            response.write(buffer.getvalue())
        except Exception as e_write:
            # En caso de error de escritura
            with open("excel_error.log", "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now()}] Error writing Excel: {e_write}\n")
            traceback.print_exc()
            raise

        buffer.close()
        return response

    except Exception as e:
        # Capturamos cualquier error global (openpyxl, I/O, etc.)
        traceback.print_exc()
        from django.http import JsonResponse
        return JsonResponse({
            "error": "Fallo al generar Excel",
            "detalle": str(e),
            "tipo": type(e).__name__
        }, status=500)